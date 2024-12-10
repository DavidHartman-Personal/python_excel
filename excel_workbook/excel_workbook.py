"""This is the main Excel Workbook class

Creates an Excel workbook class object.

    Attributes:
            workbook_filename (str): The full filename path to the Excel workbook
            worksheets (int):  A dictionary containing the Excel worksheets in the Excel workbook
            defined_names (str): An array of the defined names in the Excel workbook


    Methods
    -------

"""
__version__ = '2024.11'
__author__ = 'David Hartman'

import logging
import os
import sys
from openpyxl import load_workbook
import coloredlogs
from pprint import pp

#: This effectively defines the root of this project and so adding ..\, etc is not needed in config files
PROJECT_ROOT_DIR = os.path.dirname(os.path.dirname(__file__))

# Add script directory to the path to allow searching for modules
sys.path.insert(0, PROJECT_ROOT_DIR)

#: Directory that contains configuration files
CONF_DIR = os.path.join(PROJECT_ROOT_DIR, 'conf')

tab_separator = "\t"
comma_separator = ", "

# import any functions inside or outside of this module.  If no helpers are needed it can be removed from here
# as well as remove the file from the excel_workbook module.
# from . import helpers

coloredlogs.install(level=logging.INFO,
                    fmt="%(asctime)s %(hostname)s %(name)s %(filename)s line-%(lineno)d %(levelname)s - %(message)s",
                    datefmt='%H:%M:%S')

class ExcelWorkbook2:

    def __init__(self,
                 workbook_filename):
        self.workbook_filename = workbook_filename
        self.worksheets = {}
        self.defined_names = []
        if not os.path.exists(workbook_filename):
            logging.error("File not found [%s], if writing pass write_flag = True", str(workbook_filename))
            self.workbook = None
        else:
            self.workbook = load_workbook(filename=workbook_filename, data_only=True)
            logging.info("Creating Excel object based on file [%s]", str(workbook_filename))

    def get_defined_tables(self, worksheet_name=None):
        if not worksheet_name:
            logging.info("Reporting all defined tables on all worksheets")
            for ws in self.workbook.worksheets:
                logging.info("Worksheet [%s]", ws.title)
                for tbl in ws.tables.items():
                    print(str(tbl))
#                    print(" : " + tbl.displayName)
#                     print("   -  name = " + tbl.name)
#                     print("   -  type = " + tbl.tableType if isinstance(tbl.tableType, str) else 'n/a')
#                     print("   - range = " + tbl.ref)
#                     print("   - #cols = %d" % len(tbl.tableColumns))
#                     for col in tbl.tableColumns:
#                         print("     : " + col.name)
        else:
            logging.info("Reporting all defined tables for worksheet [%s]", worksheet_name)
            ws = self.workbook[worksheet_name]
            for tbl in ws.tables.items():
                print(str(tbl))
#                print(" : " + tbl.displayName)
#                 print("   -  name = " + tbl.name)
#                 print("   -  type = " + tbl.tableType if isinstance(tbl.tableType, str) else 'n/a')
#                 print("   - range = " + tbl.ref)
#                 print("   - #cols = %d" % len(tbl.tableColumns))
#                 for col in tbl.tableColumns:
#                     print("     : " + col.name)

    # def get_table_data_dictionary(self, worksheet_name, table_name):
    #     """Get the table data from the worksheet and named table.  Return in dictionary
    #
    #     Args:
    #         worksheet_name (str): The name of the Excel workbook
    #         table_name (str): The name of the table
    #
    #     """
    #     return_dictionary_list = []
    #     worksheet_table = []
    #     # Grab the 'data' from the table
    #     # rows_list = []
    #     ws = self.workbook.
    #     for tbl in ws.tables.items():
    #         #ws_table_name = tbl.name
    #         table_dictionary = self.get_table_data(ws[tbl.ref])
    #         worksheet_tables['TABLE_NAME'] = table_dictionary
    #     for row in table_data:
    #         # Get a list of all columns in each row
    #         cols = []
    #         for col in row:
    #             cols.append(col.value)
    #             rows_list.append(cols)
    #     header_columns = [col.upper() for col in rows_list[0]]
    #     for data_row in rows_list[1:]:
    #         row_dictionary = zip(header_columns, data_row)
    #         return_dictionary_list.append(dict(row_dictionary))
    #     logging.info("Table data [%s]:", str(return_dictionary_list))
    #     return return_dictionary_list

    def get_worksheets(self) -> []:
        """Returns an array of the worksheet objects in the workbook

        Returns:
            array object:
        """
        return_worksheets_list = list()
        for ws in self.workbook.worksheets:
            return_worksheets_list.append(ws.title)
            pp(ws)
            # pp(ws.tables.items())
            logging.info("Table data dictionary [%s]:", str(ws.tables.items()))
            # logging.debug("Looking at worksheet: [%s]", str(ws))
            # let's add it to the dictionary of worksheets if we haven't already
            if not self.worksheets.get(ws.title):
                worksheet_object = {}
                worksheet_title = ws.title
                worksheet_tables = {}
                # for tbl in ws.tables.items():
                #     #ws_table_name = tbl.name
                #     table_dictionary = self.get_table_data(ws[tbl.ref])
                #     worksheet_tables['TABLE_NAME'] = table_dictionary
                worksheet_object['WORKSHEET_NAME'] = worksheet_title
                self.worksheets[ws.title] = worksheet_object
        # logging.debug("Worksheets [%s]:", str(return_worksheets_list))
        return return_worksheets_list

    def update_defined_tables(self):
        """This function will get all defined tables on all worksheets and update the worksheets dictionary

        """
        logging.info("Gathering all defined tables in workbook [%s]", self.workbook_filename)
        for ws in self.workbook.worksheets:
            logging.info("Gathering tables from worksheet [%s]", ws.title)

    def add_worksheet_definitions(self):
        for ws in self.workbook.worksheets:
            worksheet_title = ws.title
            worksheet_definition = {'WORKSHEET_NAME': worksheet_title,
                                    'WORKSHEET_TABLES': []
                                    }
            # self.worksheets[worksheet_title] = {}
            worksheet_table = {}
            for tbl in ws._tables:
                table_name = tbl.name
                table_data = ws[tbl.ref]
                # table_data_dictionary = self.get_table_data(table_data)
                # let's add it to the dictionary of worksheets if we haven't already
                if not self.worksheets.get(ws.title):
                    self.worksheets[ws.title] = ws


def get_table_data(table_data):
    """ Get a Excel table from a worksheet and return it as a dictionary object.

    Args:
        table_data:

    """
    return_dictionary_list = []
    worksheet_table = []
    # Grab the 'data' from the table
    rows_list = []
    for row in table_data:
        # Get a list of all columns in each row
        cols = []
        for col in row:
            cols.append(col.value)
            rows_list.append(cols)
    header_columns = [col.upper() for col in rows_list[0]]
    for data_row in rows_list[1:]:
        row_dictionary = zip(header_columns, data_row)
        return_dictionary_list.append(dict(row_dictionary))
    logging.info("Table data dictionary [%s]:", str(return_dictionary_list))
    return return_dictionary_list


class ExcelWorkbook:

    def __init__(self,
                 workbook_filename):
        """

        @param workbook_filename:
        """
        self.workbook_filename = workbook_filename
        self.worksheets = {}
        self.defined_names = []
        if not os.path.exists(workbook_filename):
            logging.error("File not found [%s], if writing pass write_flag = True", str(workbook_filename))
            self.workbook = None
        else:
            self.workbook = load_workbook(filename=workbook_filename, data_only=True)
            logging.info("Creating Excel object based on file [%s]", str(workbook_filename))

    def get_worksheets(self) -> []:
        """Returns an array of the worksheet objects in the workbook and populates/updates ExcelWorkbook.worksheets array

        Returns:
            array object:
        """
        return_worksheets_list = list()
        for ws in self.workbook.worksheets:
            return_worksheets_list.append(ws.title)
            pp(ws)
            # pp(ws.tables.items())
            logging.info("Table data dictionary [%s]:", str(ws.tables.items()))
            # logging.debug("Looking at worksheet: [%s]", str(ws))
            # let's add it to the dictionary of worksheets if we haven't already
            if not self.worksheets.get(ws.title):
                worksheet_object = {}
                worksheet_title = ws.title
                worksheet_tables = {}
                # for tbl in ws.tables.items():
                #     #ws_table_name = tbl.name
                #     table_dictionary = self.get_table_data(ws[tbl.ref])
                #     worksheet_tables['TABLE_NAME'] = table_dictionary
                worksheet_object['WORKSHEET_NAME'] = worksheet_title
                self.worksheets[ws.title] = worksheet_object
        # logging.debug("Worksheets [%s]:", str(return_worksheets_list))
        return return_worksheets_list

